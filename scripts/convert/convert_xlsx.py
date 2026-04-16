#!/usr/bin/env python3
"""convert_xlsx.py — Convert Excel (.xlsx/.xls) files to Markdown.

Usage:
    python convert_xlsx.py <input_file> <output_md> [--images-dir <dir>]

Arguments:
    input_file    Path to an .xlsx or .xls file.
    output_md     Path to write the resulting Markdown file.
    --images-dir  Optional directory for extracted images.
                  Defaults to the same directory as output_md.

Exit codes:
    0   Success
    1   Failure (error details printed to stderr)

Stdout:
    A JSON summary, e.g.
    {"success": true, "file_type": "xlsx", "sheets": 3, "rows": 150, "images": 2}

Dependencies:
    openpyxl   — for .xlsx parsing
    xlrd       — for .xls parsing
    Pillow     — for image validation
"""

from __future__ import annotations

import argparse
import io
import json
import logging
import os
import re
import sys
import uuid
import zipfile
from datetime import date, datetime, time
from pathlib import Path
from typing import Dict, List, Optional, Tuple

# ---------------------------------------------------------------------------
# Optional imports — fail gracefully so we can report a clear error message
# ---------------------------------------------------------------------------
try:
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.utils.datetime import from_excel
except ImportError:
    openpyxl = None
    load_workbook = None
    from_excel = None

try:
    import xlrd
    from xlrd import xldate_as_datetime
except ImportError:
    xlrd = None
    xldate_as_datetime = None

try:
    from PIL import Image
except ImportError:
    Image = None

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Date / cell helpers  (ported from enhanced_parser.py)
# ---------------------------------------------------------------------------

def format_dt(dt: date | datetime) -> str:
    """Format a date/datetime as a readable string."""
    if isinstance(dt, datetime):
        if dt.time() == time(0, 0):
            return dt.date().isoformat()
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(dt, date):
        return dt.isoformat()
    return str(dt)


def cell_to_text(cell, wb_epoch) -> str:
    """Convert an openpyxl cell to text, fixing dates shown as numbers."""
    v = cell.value
    if v is None:
        return ""

    # Cell explicitly marked as a date
    if cell.is_date:
        if isinstance(v, (int, float)):
            try:
                v = from_excel(v, wb_epoch)
            except Exception:
                pass
        return format_dt(v)

    # Numeric value whose format looks like a date/time
    if isinstance(v, (int, float)):
        fmt = (cell.number_format or "").lower()
        if any(tok in fmt for tok in ("yy", "dd", "mm", "hh", "ss")):
            try:
                dt = from_excel(v, wb_epoch)
                return format_dt(dt)
            except Exception:
                pass

        # Fallback: integers in the serial-number date range (~1955-2050)
        if float(v).is_integer() and 20000 <= v <= 60000:
            try:
                dt = from_excel(v, wb_epoch)
                if isinstance(dt, (date, datetime)):
                    d = dt.date() if isinstance(dt, datetime) else dt
                    if date(1990, 1, 1) <= d <= date(2100, 12, 31):
                        return format_dt(dt)
            except Exception:
                pass

    return str(v)


def xls_cell_to_text(cell, datemode: int) -> str:
    """Convert an xlrd cell to text, supporting date conversion."""
    if cell.value is None or cell.value == "":
        return ""
    # ctype == 3 means XL_CELL_DATE in xlrd
    if cell.ctype == 3:
        try:
            dt = xldate_as_datetime(cell.value, datemode)
            return format_dt(dt)
        except Exception:
            pass
    return str(cell.value) if cell.value is not None else ""


# ---------------------------------------------------------------------------
# Markdown table builder  (ported from enhanced_parser.py)
# ---------------------------------------------------------------------------

def convert_list_to_markdown_table(table_data: List[List[str]]) -> List[str]:
    """Convert a 2-D list of strings into Markdown table lines."""
    if not table_data:
        return []

    markdown_table: List[str] = []
    try:
        max_cols = max(len(row) for row in table_data)
        normalized: List[List[str]] = []
        for row in table_data:
            padded = row + [""] * (max_cols - len(row))
            cleaned = [
                str(c).replace("\n", "<br>").replace("|", "\\|")
                for c in padded
            ]
            normalized.append(cleaned)

        if normalized:
            headers = (
                normalized[0]
                if len(normalized) > 1
                else [f"Col{i+1}" for i in range(max_cols)]
            )
            markdown_table.append("| " + " | ".join(headers) + " |")
            markdown_table.append("| " + " | ".join(["---"] * len(headers)) + " |")

            data_rows = normalized[1:] if len(normalized) > 1 else normalized
            for row in data_rows:
                markdown_table.append("| " + " | ".join(row) + " |")
    except Exception as exc:
        logger.warning("Table conversion failed: %s", exc)
        return []

    return markdown_table


# ---------------------------------------------------------------------------
# Markdown cleanup
# ---------------------------------------------------------------------------

def clean_markdown(lines: List[str]) -> str:
    """Join lines and normalise whitespace."""
    content = "\n".join(lines)
    content = re.sub(r"\n{3,}", "\n\n", content)
    result = [l.rstrip() for l in content.split("\n")]
    while result and not result[0].strip():
        result.pop(0)
    while result and not result[-1].strip():
        result.pop()
    return "\n".join(result)


# ---------------------------------------------------------------------------
# Image helpers
# ---------------------------------------------------------------------------

IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp"}


def validate_image(path: Path) -> bool:
    """Return True if *path* is a valid image readable by Pillow."""
    if Image is None:
        return True  # can't validate without Pillow — assume OK
    try:
        with Image.open(path) as img:
            img.verify()
        return True
    except Exception:
        return False


def extract_xlsx_images(file_path: str, images_dir: Path) -> List[Dict]:
    """Extract embedded images from an .xlsx file (which is a ZIP)."""
    images_info: List[Dict] = []
    try:
        with zipfile.ZipFile(file_path, "r") as zf:
            media_files = [
                n for n in zf.namelist()
                if n.startswith("xl/media/")
                and Path(n).suffix.lower() in IMAGE_EXTENSIONS
            ]
            logger.info("Found %d media files in XLSX", len(media_files))

            for media_file in media_files:
                try:
                    data = zf.read(media_file)
                    if len(data) < 100:
                        continue

                    ext = Path(media_file).suffix.lower().lstrip(".")
                    fname = f"{uuid.uuid4()}.{ext}"
                    dest = images_dir / fname

                    dest.parent.mkdir(parents=True, exist_ok=True)
                    dest.write_bytes(data)

                    if validate_image(dest):
                        images_info.append({
                            "filename": fname,
                            "path": str(dest),
                            "size": len(data),
                            "format": ext,
                            "source": media_file,
                        })
                        logger.info("Extracted image: %s", media_file)
                    else:
                        dest.unlink(missing_ok=True)
                        logger.warning("Invalid image removed: %s", media_file)
                except Exception as exc:
                    logger.warning("Error processing image %s: %s", media_file, exc)
    except Exception as exc:
        logger.error("XLSX image extraction failed: %s", exc)

    return images_info


# ---------------------------------------------------------------------------
# Core parsers
# ---------------------------------------------------------------------------

def parse_xlsx(file_path: str, images_dir: Path) -> Dict:
    """Parse an .xlsx file and return a result dict."""
    if openpyxl is None or load_workbook is None or from_excel is None:
        raise ImportError("openpyxl is required to parse .xlsx files — pip install openpyxl")

    workbook = load_workbook(filename=file_path, data_only=True)
    wb_epoch = workbook.epoch
    markdown_lines: List[str] = []
    total_rows = 0

    # Extract images
    images_info = extract_xlsx_images(file_path, images_dir)

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Skip empty sheets
        if sheet.max_row is None or (sheet.max_row <= 1 and sheet.max_column <= 1):
            continue

        table_data: List[List[str]] = []
        for row in sheet.iter_rows(
            min_row=1, max_row=sheet.max_row,
            min_col=1, max_col=sheet.max_column,
        ):
            row_values = [cell_to_text(c, wb_epoch) for c in row]
            if any(v.strip() for v in row_values):
                table_data.append(row_values)

        if not table_data:
            continue

        total_rows += len(table_data)

        markdown_lines.append(f"## Sheet: {sheet_name}")
        markdown_lines.append("")
        markdown_lines.extend(convert_list_to_markdown_table(table_data))
        markdown_lines.append("")

        # Append images that belong to this sheet (if mapping exists)
        sheet_images = [i for i in images_info if i.get("sheet") == sheet_name]
        for img in sheet_images:
            rel = os.path.relpath(img["path"], images_dir.parent).replace("\\", "/")
            markdown_lines.append(f"![{img['filename']}]({rel})")
        if sheet_images:
            markdown_lines.append("")

    # Images that weren't assigned to a specific sheet — append at the end
    unassigned = [i for i in images_info if "sheet" not in i or i.get("sheet") is None]
    if unassigned:
        markdown_lines.append("## Images")
        markdown_lines.append("")
        for img in unassigned:
            rel = os.path.relpath(img["path"], images_dir.parent).replace("\\", "/")
            markdown_lines.append(f"![{img['filename']}]({rel})")
        markdown_lines.append("")

    md_text = clean_markdown(markdown_lines)

    return {
        "success": True,
        "file_type": "xlsx",
        "sheets": len(workbook.sheetnames),
        "rows": total_rows,
        "images": len(images_info),
        "markdown": md_text,
    }


def parse_xls(file_path: str, images_dir: Path) -> Dict:
    """Parse an .xls file and return a result dict."""
    if xlrd is None or xldate_as_datetime is None:
        raise ImportError("xlrd is required to parse .xls files — pip install xlrd")

    workbook = xlrd.open_workbook(file_path, formatting_info=False)
    datemode = workbook.datemode
    markdown_lines: List[str] = []
    total_rows = 0

    for sheet_idx in range(workbook.nsheets):
        sheet = workbook.sheet_by_index(sheet_idx)
        sheet_name = workbook.sheet_names()[sheet_idx]

        if sheet.nrows == 0:
            continue

        table_data: List[List[str]] = []
        for row_idx in range(sheet.nrows):
            row_data = [
                xls_cell_to_text(sheet.cell(row_idx, col_idx), datemode)
                for col_idx in range(sheet.ncols)
            ]
            if any(v.strip() for v in row_data):
                table_data.append(row_data)

        if not table_data:
            continue

        total_rows += len(table_data)

        markdown_lines.append(f"## Sheet: {sheet_name}")
        markdown_lines.append("")
        markdown_lines.extend(convert_list_to_markdown_table(table_data))
        markdown_lines.append("")

    md_text = clean_markdown(markdown_lines)

    return {
        "success": True,
        "file_type": "xls",
        "sheets": workbook.nsheets,
        "rows": total_rows,
        "images": 0,  # xls image extraction is very limited
        "markdown": md_text,
    }


# ---------------------------------------------------------------------------
# CLI entry-point
# ---------------------------------------------------------------------------

def main() -> int:
    parser = argparse.ArgumentParser(
        description="Convert an Excel file (.xlsx / .xls) to Markdown.",
    )
    parser.add_argument("input_file", help="Path to the .xlsx or .xls file")
    parser.add_argument("output_md", help="Path to write the resulting .md file")
    parser.add_argument(
        "--images-dir",
        default=None,
        help="Directory to save extracted images (default: same dir as output_md)",
    )
    args = parser.parse_args()

    input_path = Path(args.input_file).resolve()
    output_path = Path(args.output_md).resolve()
    images_dir = Path(args.images_dir).resolve() if args.images_dir else output_path.parent

    # Validate input
    if not input_path.exists():
        summary = {"success": False, "error": f"Input file not found: {input_path}"}
        print(json.dumps(summary))
        return 1

    ext = input_path.suffix.lower()
    if ext not in (".xlsx", ".xls"):
        summary = {"success": False, "error": f"Unsupported file extension: {ext}"}
        print(json.dumps(summary))
        return 1

    # Ensure output & images directories exist
    output_path.parent.mkdir(parents=True, exist_ok=True)
    images_dir.mkdir(parents=True, exist_ok=True)

    try:
        if ext == ".xlsx":
            result = parse_xlsx(str(input_path), images_dir)
        else:
            result = parse_xls(str(input_path), images_dir)

        # Write markdown
        output_path.write_text(result["markdown"], encoding="utf-8")

        # Print JSON summary to stdout
        summary = {
            "success": True,
            "file_type": result["file_type"],
            "sheets": result["sheets"],
            "rows": result["rows"],
            "images": result["images"],
        }
        print(json.dumps(summary))
        return 0

    except Exception as exc:
        logger.error("Conversion failed: %s", exc, exc_info=True)
        summary = {"success": False, "error": str(exc)}
        print(json.dumps(summary))
        return 1


if __name__ == "__main__":
    sys.exit(main())
